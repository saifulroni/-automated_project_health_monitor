import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
from rag_engine import run_rag_engine

#rag colors
RAG_FILLS = {
    'Red':   PatternFill("solid", fgColor="FF4444"),
    'Amber': PatternFill("solid", fgColor="FFA500"),
    'Green': PatternFill("solid", fgColor="44BB44"),
    'Grey':  PatternFill("solid", fgColor="AAAAAA"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

def style_header(cell, text):
    cell.value = text
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')

def add_thin_border(cell):
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_portfolio_summary(wb, df, project_rags):
    ws = wb.create_sheet("Portfolio_Summary")
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    style_header(ws['A1'], "LEMU TAJIN LAUNCH — PORTFOLIO HEALTH SUMMARY")
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"Report generated: {datetime.today().strftime('%A, %d %B %Y')}"
    ws['A2'].font = Font(italic=True, size=9)

    headers = ['Project', 'Overall RAG', 'Total Tasks', '% Complete', 'Overdue Tasks']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        add_thin_border(cell)

    for row_idx, (project, rag) in enumerate(project_rags.items(), 5):
        proj_df = df[df['Project_Name'] == project]
        total = len(proj_df)
        pct_done = round(proj_df['Pct_Complete'].mean(), 1)
        overdue = len(proj_df[(proj_df['RAG'] == 'Red') & (proj_df['Status'] != 'Completed')])

        ws.cell(row=row_idx, column=1, value=project)
        rag_cell = ws.cell(row=row_idx, column=2, value=rag)
        rag_cell.fill = RAG_FILLS.get(rag, RAG_FILLS['Grey'])
        rag_cell.font = Font(bold=True, color="FFFFFF")
        rag_cell.alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=3, value=total).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=4, value=f"{pct_done}%").alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=5, value=overdue).alignment = Alignment(horizontal='center')

        for col in range(1, 6):
            add_thin_border(ws.cell(row=row_idx, column=col))

def write_project_dashboard(wb, df, project_name, project_rag, critical_path):
    sheet_name = project_name.replace(' ', '_')[:30]
    ws = wb.create_sheet(sheet_name)

    
    ws.merge_cells('A1:H1')
    ws['A1'] = f"PROJECT HEALTH DASHBOARD — {project_name.upper()}"
    ws['A1'].fill = HEADER_FILL
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=13)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"Week of {datetime.today().strftime('%d %B %Y')}   |   Overall Status:"
    ws['A2'].font = Font(size=10)
    rag_cell = ws['E2']
    rag_cell.value = project_rag
    rag_cell.fill = RAG_FILLS.get(project_rag, RAG_FILLS['Grey'])
    rag_cell.font = Font(bold=True, color="FFFFFF", size=12)
    rag_cell.alignment = Alignment(horizontal='center')

    
    proj_df = df[df['Project_Name'] == project_name]
    total = len(proj_df)
    completed = len(proj_df[proj_df['Status'] == 'Completed'])
    overdue = len(proj_df[(proj_df['RAG'] == 'Red') & (proj_df['Status'] != 'Completed')])
    blocked = len(proj_df[proj_df['Status'] == 'Blocked'])
    avg_pct = round(proj_df['Pct_Complete'].mean(), 1)

    kpis = [
        ('Total Tasks', total),
        ('Completed', completed),
        ('Avg % Complete', f"{avg_pct}%"),
        ('Overdue', overdue),
        ('Blocked', blocked),
    ]
    for col, (label, value) in enumerate(kpis, 1):
        ws.cell(row=4, column=col, value=label).font = Font(bold=True, size=9)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
        val_cell = ws.cell(row=5, column=col, value=value)
        val_cell.font = Font(bold=True, size=14)
        val_cell.alignment = Alignment(horizontal='center')

    
    headers = ['Task ID', 'Task Name', 'Owner', 'Start', 'End Date',
               'Status', '% Done', 'RAG', 'Critical?', 'Comments']
    col_widths = [9, 38, 12, 12, 12, 14, 8, 8, 10, 40]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=7, column=col, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        add_thin_border(cell)

    for row_idx, (_, task) in enumerate(proj_df.iterrows(), 8):
        data = [
            task['Task_ID'],
            task['Task_Name'],
            task['Owner'],
            task['Start_Date'].strftime('%d/%m/%y') if pd.notna(task['Start_Date']) else '',
            task['Planned_End_Date'].strftime('%d/%m/%y') if pd.notna(task['Planned_End_Date']) else '',
            task['Status'],
            f"{int(task['Pct_Complete'])}%",
            task['RAG'],
            '★ YES' if task['Is_Critical'] else '',
            str(task['Comments']) if pd.notna(task['Comments']) else '',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            add_thin_border(cell)
            if col == 8:  # RAG column
                cell.fill = RAG_FILLS.get(str(value), RAG_FILLS['Grey'])
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')

    
    flagged = proj_df[proj_df['RAG'].isin(['Red', 'Amber']) & (proj_df['Status'] != 'Completed')]
    start_row = 8 + len(proj_df) + 2
    ws.cell(row=start_row, column=1, value="⚠ ATTENTION REQUIRED — OVERDUE & BLOCKED TASKS")
    ws.cell(row=start_row, column=1).font = Font(bold=True, color="FF0000", size=11)
    ws.merge_cells(f'A{start_row}:J{start_row}')

    for row_idx, (_, task) in enumerate(flagged.iterrows(), start_row + 1):
        today = pd.Timestamp(datetime.today().date())
        days_late = (today - task['Planned_End_Date']).days if pd.notna(task['Planned_End_Date']) else 0
        ws.cell(row=row_idx, column=1, value=task['Task_ID'])
        ws.cell(row=row_idx, column=2, value=task['Task_Name'])
        ws.cell(row=row_idx, column=3, value=task['Owner'])
        ws.cell(row=row_idx, column=4, value=f"{max(days_late, 0)}d late" if days_late > 0 else task['Status'])
        ws.cell(row=row_idx, column=5, value=str(task['Comments']) if pd.notna(task['Comments']) else '')
        rag_cell = ws.cell(row=row_idx, column=8, value=task['RAG'])
        rag_cell.fill = RAG_FILLS.get(task['RAG'], RAG_FILLS['Grey'])
        rag_cell.font = Font(bold=True, color="FFFFFF")
        for col in [1, 2, 3, 4, 5, 8]:
            add_thin_border(ws.cell(row=row_idx, column=col))

def generate_dashboard(tasks_filepath, output_filepath):
    df, project_rags, critical_path = run_rag_engine(tasks_filepath)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  

    write_portfolio_summary(wb, df, project_rags)
    for project in df['Project_Name'].unique():
        write_project_dashboard(wb, df, project, project_rags[project], critical_path)

    wb.save(output_filepath)
    print(f"Dashboard saved: {output_filepath}")

if __name__ == '__main__':
    generate_dashboard('Tasks.xlsx', 'outputs/lemu_tajin_dashboard.xlsx')